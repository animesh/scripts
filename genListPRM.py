#python3 genListPRM.py
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings("ignore")

# ── 1. RELOAD LFQ DATA ────────────────────────────────────────────────────────
ISNS_FILE = "/home/animeshs/promec/promec/TIMSTOF/LARS/2025/250805_Kamila/DIANNv2p2/report rq.ha..gg_matrix.tsv4118ISNS0.10.50.1BioRemGroups.txt4LFQvsntTestBH.xlsx"
ITNT_FILE = "/home/animeshs/promec/promec/TIMSTOF/LARS/2025/250805_Kamila/DIANNv2p2/report rq.ha..gg_matrix.tsv4118ITNT0.10.50.1BioRemGroups.txt4LFQvsntTestBH.xlsx"
isns_lfq = pd.read_excel(ISNS_FILE)
itnt_lfq = pd.read_excel(ITNT_FILE)
IPA_SECTIONS = {
    "ISNS": ("/home/animeshs/promec/promec/TIMSTOF/LARS/2025/250805_Kamila/DIANNv2p2/ISNS.xls",
             dict(pathways=(22,931), upstream=(931,2538), molecules=(5432,5924))),
    "ITNT": ("/home/animeshs/promec/promec/TIMSTOF/LARS/2025/250805_Kamila/DIANNv2p2/ITNT.xls",
             dict(pathways=(22,728), upstream=(728,1330), molecules=(3910,4141))),
}

OUTPUT = "/home/animeshs/promec/promec/TIMSTOF/LARS/2025/250805_Kamila/BCC_PRM_Panel_v2_IPA.xlsx"

is_cols = [c for c in isns_lfq.columns if c.startswith("IS")]
ns_cols = [c for c in isns_lfq.columns if c.startswith("NS")]
it_cols = [c for c in itnt_lfq.columns if c.startswith("IT")]
nt_cols = [c for c in itnt_lfq.columns if c.startswith("NT")]
isns_lfq["n_IS"] = isns_lfq[is_cols].notna().sum(axis=1)
isns_lfq["n_NS"] = isns_lfq[ns_cols].notna().sum(axis=1)
itnt_lfq["n_IT"] = itnt_lfq[it_cols].notna().sum(axis=1)
itnt_lfq["n_NT"] = itnt_lfq[nt_cols].notna().sum(axis=1)

# ── 2. PARSE IPA XLS FILES ────────────────────────────────────────────────────
def read_ipa_section(fname, start_row, end_row):
    df = pd.read_excel(fname, engine="xlrd", header=None,
                       skiprows=start_row+1, nrows=end_row-start_row-2)
    df.columns = df.iloc[0]
    return df.iloc[1:].reset_index(drop=True)

ipa = {}
for label, (fname, secs) in IPA_SECTIONS.items():
    d = {}
    for key, (s, e) in secs.items():
        df = read_ipa_section(fname, s, e)
        df.columns = [str(c).strip() for c in df.columns]
        d[key] = df
    ipa[label] = d

# ── 3. EXTRACT SIGNIFICANT PATHWAYS (p<0.05, i.e. -log10 > 1.3) ──────────────
def get_sig_pathways(label):
    df = ipa[label]["pathways"].copy()
    lp_col = "-log(p-value)"
    df[lp_col] = pd.to_numeric(df[lp_col], errors="coerce")
    zs_col = "zScore"
    df[zs_col] = pd.to_numeric(df[zs_col], errors="coerce")
    sig = df[df[lp_col] > 1.3].sort_values(lp_col, ascending=False).copy()
    sig["p_value"] = 10 ** (-sig[lp_col])
    sig["label"] = label
    return sig[["label","Ingenuity Canonical Pathways", lp_col, zs_col, "Ratio","Molecules","p_value"]]

isns_paths = get_sig_pathways("ISNS")
itnt_paths = get_sig_pathways("ITNT")
all_paths  = pd.concat([isns_paths, itnt_paths], ignore_index=True)

# ── 4. EXTRACT SIGNIFICANT UPSTREAM REGULATORS ───────────────────────────────
def get_sig_upstream(label):
    df = ipa[label]["upstream"].copy()
    pc = "p-value of overlap"
    df[pc] = pd.to_numeric(df[pc], errors="coerce")
    az = "Activation z-score"
    df[az] = pd.to_numeric(df[az], errors="coerce")
    sig = df[df[pc] < 0.001].sort_values(pc).copy()
    sig["label"] = label
    keep = ["label","Upstream Regulator","Expr Log Ratio","Molecule Type",
            "Predicted Activation State","Activation z-score",
            "Bias-corrected z-score","p-value of overlap",
            "Target molecules in dataset"]
    return sig[[c for c in keep if c in sig.columns]]

isns_ups = get_sig_upstream("ISNS")
itnt_ups = get_sig_upstream("ITNT")

# ── 5. EXTRACT IPA ANALYSIS-READY MOLECULES (genes in the IPA analysis) ──────
def get_ipa_molecules(label):
    df = ipa[label]["molecules"].copy()
    fc_col = "Expr Log Ratio"
    qv_col = "Expr False Discovery Rate (q-value)"
    df[fc_col] = pd.to_numeric(df[fc_col], errors="coerce")
    df[qv_col] = pd.to_numeric(df[qv_col], errors="coerce")
    return df[["Symbol",fc_col,qv_col,"Expr p-value"]].dropna(subset=["Symbol"])

isns_mols = get_ipa_molecules("ISNS")
itnt_mols = get_ipa_molecules("ITNT")
isns_mols.columns = ["Gene","FC_IS_NS_IPA","BH_IS_NS_IPA","Pval_IS_NS_IPA"]
itnt_mols.columns = ["Gene","FC_IT_NT_IPA","BH_IT_NT_IPA","Pval_IT_NT_IPA"]

# ── 6. BUILD GENE→PATHWAY AND GENE→REGULATOR ANNOTATIONS ─────────────────────
def genes_in_pathways(paths_df, label):
    """Return dict gene→set of pathways (for genes appearing in any pathway molecule list)."""
    g2p = {}
    for _, row in paths_df[paths_df["label"]==label].iterrows():
        mols = str(row["Molecules"])
        pw   = str(row["Ingenuity Canonical Pathways"])
        for g in mols.split(","):
            g = g.strip()
            if g and g != "nan":
                g2p.setdefault(g, set()).add(pw)
    return g2p

isns_g2p = genes_in_pathways(all_paths, "ISNS")
itnt_g2p = genes_in_pathways(all_paths, "ITNT")

def regulators_of(gene, ups_df, label):
    """Return list of (regulator, activation_state, z_score) for a gene."""
    results = []
    for _, row in ups_df[ups_df["label"]==label].iterrows():
        targets = str(row.get("Target molecules in dataset",""))
        reg = str(row["Upstream Regulator"])
        if gene in [t.strip() for t in targets.split(",")]:
            results.append((reg,
                             row.get("Predicted Activation State",""),
                             row.get("Activation z-score","")))
    return results

# ── 7. KEY IPA-DERIVED ADDITIONS TO PANEL ─────────────────────────────────────
# A. Keratinization pathway is significantly DOWN in IT vs NT (z=-2.12)
#    KRT5, KRT14, KRT17, KRT20, TGM5 — important nodular BCC markers
# B. Elastic fibre formation UP in IT (FBLN2, FBLN5, EMILIN2, FN1)
# C. TGFB1 is confirmed ACTIVATED upstream regulator in IS vs NS stroma
# D. TP53 ACTIVATED in IT vs NT
# E. New proteins in IPA but not in previous seed:
IPA_ADDITIONS = {
    "Keratinocyte Differentiation / BCC Identity": (9, [
        "KRT5","KRT14","KRT17","KRT20","KRT75","KRT76","KRT80","TGM5",
        "ITGB4","FERMT1","FERMT2","DSG1","DSG3","DSP","EVPL"
    ]),
    "Elastic Fibre / Basement Membrane": (8, [
        "FBLN2","FBLN5","EMILIN2","ELN","MFAP2","MFAP4","LTBP2","LTBP4","FBN2"
    ]),
    "mRNA Processing / Transcription (ITNT)": (4, [
        "SF3A2","SF3A3","SF3B3","SF3B6","POLR2E","POLR3A","POLR3B",
        "HNRNPU","HNRNPR","CDK9","EP300","CHD4"
    ]),
    "Mitochondrial Ribosome / Translation (ITNT)": (3, [
        "MRPS17","MRPS16","MRPS23","MRPS27","MRPS35","MRPL13","MRPL15",
        "MRPL19","MRPL38","MRPL39","MRPL44"
    ]),
}

# ── 8. MERGED BIOLOGICAL CATEGORIES (v1 + IPA additions) ─────────────────────
CATEGORIES = {
    "Type-2 Immunity": (10, [
        "POSTN","ARG1","MRC1","CCL17","CCL22","IL4","IL13","IL5","IL33","TSLP",
        "CXCL12","CLCA1","PDGFRA","IL4R","STAT6","GATA3","IL2RA",
        "CHI3L1","CHI3L2","CHIT1"
    ]),
    "Type-1 Immunity": (9, [
        "IFNG","CXCL10","CXCL9","CXCL11","STAT1","IRF1","IRF7","IFIT1","IFIT3",
        "GBP1","GBP2","MX1","PSMB8","TAP1","B2M","CD8A","GZMB","PRF1",
    ]),
    "Type-3 / Inflammasome": (8, [
        "S100A8","S100A9","S100A12","S100A6","NLRP3","CASP1","IL1B","IL17A",
        "IL6","CXCL1","CXCL2","MMP8","MMP9","ELANE","MPO","LTF"
    ]),
    "TGFb / Immunosuppression": (10, [
        "TGFB1","TGFB2","TGFB3","LTBP2","LTBP4","TGFBI","SMAD2","SMAD3",
        "CXCL12","FOXP3","PDCD1","CD274","IDO1","CD163","MRC1","ARG1",
        "SERPINE2","CCN2"
    ]),
    "CAF Program": (10, [
        "FAP","ACTA2","PDGFRB","PDGFRA","PDPN","S100A4","VIM","THY1",
        "LRRC15","CXCL12","IGFBP3","IGFBP5","IGFBP7",
        "COL11A1","COL10A1","MMP11","INHBA","SFRP2","SFRP4","WNT5A"
    ]),
    "ECM Remodeling": (9, [
        "COL1A1","COL1A2","COL3A1","COL4A1","COL4A2","COL5A1","COL6A3",
        "COL10A1","COL11A1","COL12A1","COL14A1","COL17A1","COL26A1",
        "FN1","VCAN","TNC","POSTN","THBS1","THBS2","THBS4",
        "MMP2","MMP12","MMP14","MMP11","MMP1",
        "LOXL1","LOXL2","LOXL3","LOX",
        "FBLN1","FBLN2","FBLN5","ELN","EMILIN2","LTBP2","FBN2",
        "LAMB2","HSPG2","BGN","DCN","LUM","ASPN","MATN2","TNXB","HAPLN3",
        "PXDN","COMP"
    ]),
    "Elastic Fibre / Basement Membrane": (8, [
        "FBLN2","FBLN5","EMILIN2","ELN","MFAP2","MFAP4","LTBP2","LTBP4","FBN2",
        "LAMB2","HSPG2","ITGA3","ITGA6","ITGB4","ITGB5","ITGB6"
    ]),
    "Hedgehog Signaling": (8, [
        "SHH","IHH","DHH","PTCH1","PTCH2","SMO","GLI1","GLI2","GLI3",
        "HHIP","BOC","GAS1","CDON","SUFU","WNT5A","SFRP2","SFRP4"
    ]),
    "Angiogenesis / Vasculature": (7, [
        "VEGFA","VEGFB","VEGFC","KDR","FLT1","PECAM1","CDH5",
        "ANGPT1","ANGPT2","THBS1","THBS2","COL18A1","MCAM"
    ]),
    "Keratinocyte Differentiation / BCC Identity": (9, [
        "KRT5","KRT14","KRT17","KRT20","KRT75","KRT76","KRT80","TGM5",
        "TGM2","ITGB4","FERMT1","FERMT2","DSG1","DSG3","EVPL"
    ]),
    "Complement / Innate": (6, [
        "C1QA","C1QB","C1QC","C3","C4A","C4B","CFB","CFH","CFI",
        "FCN1","FCN2","FCN3","C1S","MBL2","PRTN3","ELANE"
    ]),
    "High-FC Infiltrative Biomarker": (7, [
        "FAT2","CRACD","CLEC16A","CLK1","FAM76B","TARS3","SIPA1L3",
        "CHIT1","CDH5","TGM2","THBS4","FBLN2","FCN3","FAP",
        "LRRC15","DDR2","MMP1","PADI2","APCDD1L","PECAM1"
    ]),
    "High-FC Nodular Biomarker": (7, [
        "ALDH1A2","TAGLN3","TUBB8","ADSS1",
        "TYRP1","ART3","FAT4","LFNG","NAALAD2","EPHX3","RBP1",
        "SERPINA12","SLC27A6","SCGB1D2","HAL","MBP"
    ]),
    "Tumor Markers (BCC)": (7, [
        "KRT5","KRT14","KRT17","KRT6A","KRT6B","TP53","CDKN2A",
        "PTCH1","SHH","SMO","GLI1","GLI2","TGM2",
        "MKI67","PCNA","CDK4","CDK6","CCND1"
    ]),
    "Lipid / Metabolic Reprogramming": (5, [
        "ALDH1A2","ALDH1A3","ALDH3A1","RBP1","FABP4","FABP5",
        "ACSL3","ACSL4","FASN","NNMT","SCD","HMGCR",
        "SLC27A6","SLC27A1","ADIPOQ","IDH1","GLS"
    ]),
    "mRNA Processing / Transcription (ITNT)": (4, [
        "SF3A2","SF3A3","SF3B3","SF3B6","POLR2E","POLR3A","POLR3B",
        "HNRNPU","HNRNPR","CDK9","EP300","CHD4","ARID1A","ARID2"
    ]),
    "Mitochondrial (ITNT)": (3, [
        "MRPS17","MRPS16","MRPS23","MRPS27","MRPS35","MRPL13","MRPL15",
        "MRPL19","MRPL38","MRPL39","MRPL44","TOMM40","SAMM50","MFN2"
    ]),
}

def assign_categories(gene):
    cats = [c for c,(w,gs) in CATEGORIES.items() if gene in gs]
    return "; ".join(cats) if cats else "Unclassified"

def assign_priority(gene):
    ws = [w for c,(w,gs) in CATEGORIES.items() if gene in gs]
    return max(ws) if ws else 3

# ── 9. BUILD MERGED DATASET ───────────────────────────────────────────────────
def prep(df, n1, n2):
    n1_max = df[n1].max(); n2_max = df[n2].max()
    def tier(r):
        bh = r["CorrectedPValueBH"]
        n1v,n2v = r[n1], r[n2]
        if pd.isna(bh):
            mx = max(n1v,n2v)
            if mx>=3 and min(n1v,n2v)==0: return "Present/Absent (n≥3)"
            if mx>=2:                      return "Present/Absent (n=1-2)"
            return "Singleton"
        if bh<0.05: return "BH<0.05"
        if bh<0.20: return "BH<0.20"
        return "Trend"
    df["tier"] = df.apply(tier, axis=1)
    return df

isns_lfq = prep(isns_lfq,"n_IS","n_NS")
itnt_lfq = prep(itnt_lfq,"n_IT","n_NT")

isns_s = isns_lfq[["Gene","Peptides","Log2MedianChange","CorrectedPValueBH",
                    "n_IS","n_NS","tier","grp1CV","grp2CV"]].copy()
itnt_s = itnt_lfq[["Gene","Peptides","Log2MedianChange","CorrectedPValueBH",
                    "n_IT","n_NT","tier","grp1CV","grp2CV"]].copy()
isns_s.columns = ["Gene","Peptides","FC_IS_NS","BH_IS_NS","n_IS","n_NS","Tier_IS_NS","CV_IS","CV_NS"]
itnt_s.columns = ["Gene","Peptides","FC_IT_NT","BH_IT_NT","n_IT","n_NT","Tier_IT_NT","CV_IT","CV_NT"]

merged = isns_s.merge(itnt_s, on="Gene", suffixes=("_s","_t"), how="outer")
merged["Peptides"] = merged["Peptides_s"].combine_first(merged["Peptides_t"])
merged = merged.drop(columns=["Peptides_s","Peptides_t"])

# Merge IPA molecule lists
merged = merged.merge(isns_mols, on="Gene", how="left")
merged = merged.merge(itnt_mols, on="Gene", how="left")

merged["Category"]         = merged["Gene"].apply(assign_categories)
merged["CategoryPriority"] = merged["Gene"].apply(assign_priority)

# IPA pathway annotation
merged["IPA_Pathways_ISNS"] = merged["Gene"].apply(
    lambda g: "; ".join(sorted(isns_g2p.get(g,[]))) if isns_g2p.get(g) else "")
merged["IPA_Pathways_ITNT"] = merged["Gene"].apply(
    lambda g: "; ".join(sorted(itnt_g2p.get(g,[]))) if itnt_g2p.get(g) else "")

# ── 10. SCORING ───────────────────────────────────────────────────────────────
TIER_SCORE = {"BH<0.05":5,"BH<0.20":4,"Present/Absent (n≥3)":4,
              "Present/Absent (n=1-2)":2,"Trend":1,"Singleton":0}

def score(row):
    ev  = sum(TIER_SCORE.get(str(row.get(c,"")),0) for c in ["Tier_IS_NS","Tier_IT_NT"])
    fc  = sum((3 if abs(v or 0)>=3 else 2 if abs(v or 0)>=1.5 else 1 if abs(v or 0)>=0.5 else 0)
              for v in [row.get("FC_IS_NS",0), row.get("FC_IT_NT",0)] if pd.notna(v))
    p   = row.get("Peptides",0) or 0
    pep = 4 if p>=10 else 3 if p>=5 else 2 if p>=2 else 1 if p>=1 else 0
    f1,f2 = row.get("FC_IS_NS",0) or 0, row.get("FC_IT_NT",0) or 0
    con = 2 if ((f1>0.3 and f2>0.3) or (f1<-0.3 and f2<-0.3)) else 0
    # IPA bonus: gene appears in a significant pathway
    ipa_bonus = 2 if (row.get("IPA_Pathways_ISNS") or row.get("IPA_Pathways_ITNT")) else 0
    bio = row["CategoryPriority"]
    return ev*2 + fc*2 + pep + con + ipa_bonus + bio

merged["TotalScore"] = merged.apply(score, axis=1)

# ── 11. SEED + INCLUSION CRITERIA ─────────────────────────────────────────────
SEED_FORCED = {
    "POSTN","ARG1","TGFB1","CXCL12","CCL17","MRC1",
    "THBS1","TNC","MMP14","VCAN","COL10A1","COL11A1",
    "FAP","ACTA2","FN1","COL1A1","COL3A1","COL4A1","COL6A3",
    "S100A8","S100A9",
    "STAT1","CXCL10","CXCL9",
    "CD8A","CD163","FOXP3","GATA3",
    "GLI1","GLI2","PTCH1","SMO",
    "MMP2","MMP11","MMP12","LOXL2",
    "FBLN2","FBLN5","EMILIN2","THBS4",
    "CDH5","TGM2","ALDH1A2","RBP1",
    # IPA-derived additions
    "KRT5","KRT14","KRT17","ITGB4","LTBP2","SERPINE2","WNT5A",
    "LRRC15","MRPS17",  # top scoring from ITNT
}

def include(row):
    g = row["Gene"]
    if g in SEED_FORCED: return True
    bh1,bh2 = row.get("BH_IS_NS"), row.get("BH_IT_NT")
    if (pd.notna(bh1) and bh1<0.2) or (pd.notna(bh2) and bh2<0.2): return True
    t1,t2 = str(row.get("Tier_IS_NS","")), str(row.get("Tier_IT_NT",""))
    f1,f2 = abs(row.get("FC_IS_NS",0) or 0), abs(row.get("FC_IT_NT",0) or 0)
    if ("Present/Absent (n≥3)" in t1 and f1>5) or \
       ("Present/Absent (n≥3)" in t2 and f2>5): return True
    # IPA-boosted
    if (row.get("IPA_Pathways_ISNS") or row.get("IPA_Pathways_ITNT")) and \
       row["CategoryPriority"] >= 7 and (row.get("Peptides",0) or 0) >= 2: return True
    if row["Category"] != "Unclassified" and (row.get("Peptides",0) or 0) >= 2 \
       and row["TotalScore"] >= 14: return True
    return False

candidates = merged[merged.apply(include, axis=1)].copy()
candidates = candidates.sort_values("TotalScore", ascending=False)
print(f"Candidate pool: {len(candidates)}")

# ── 12. TRIMMING ──────────────────────────────────────────────────────────────
TARGET = 80
CATEGORY_MIN = {
    "Type-2 Immunity":3, "Type-1 Immunity":3, "Type-3 / Inflammasome":2,
    "TGFb / Immunosuppression":3, "CAF Program":3, "ECM Remodeling":4,
    "Elastic Fibre / Basement Membrane":2,
    "Keratinocyte Differentiation / BCC Identity":3,
    "Hedgehog Signaling":2, "High-FC Infiltrative Biomarker":3,
    "High-FC Nodular Biomarker":3, "Tumor Markers (BCC)":2,
}

def get_cats(s):
    if pd.isna(s) or s=="Unclassified": return []
    return [c.strip() for c in s.split(";")]

selected = []
cat_count = {c:0 for c in CATEGORY_MIN}

def add_gene(gene, cat_str):
    selected.append(gene)
    for c in get_cats(cat_str):
        if c in cat_count: cat_count[c]+=1

# Pass 1: forced seeds
for _, row in candidates.iterrows():
    if row["Gene"] in SEED_FORCED:
        add_gene(row["Gene"], row["Category"])

# Pass 2: category minimums
for _, row in candidates.iterrows():
    if row["Gene"] in selected: continue
    if len(selected)>=TARGET: break
    for c in get_cats(row["Category"]):
        if c in CATEGORY_MIN and cat_count.get(c,0)<CATEGORY_MIN[c]:
            add_gene(row["Gene"], row["Category"]); break

# Pass 3: fill by score
for _, row in candidates.iterrows():
    if row["Gene"] in selected: continue
    if len(selected)>=TARGET: break
    t1,t2 = str(row.get("Tier_IS_NS","")), str(row.get("Tier_IT_NT",""))
    if t1=="Singleton" and (pd.isna(row.get("Tier_IT_NT")) or t2=="Singleton"): continue
    if row["Category"]=="Unclassified" and row["TotalScore"]<16: continue
    add_gene(row["Gene"], row["Category"])

print(f"Final panel: {len(selected)}")

candidates["PanelStatus"] = candidates["Gene"].apply(
    lambda g: "SELECTED" if g in selected else "TRIMMED")

def trim_reason(row):
    if row["PanelStatus"]=="SELECTED": return "Keep"
    t1,t2 = str(row.get("Tier_IS_NS","")), str(row.get("Tier_IT_NT",""))
    p = row.get("Peptides",0) or 0
    if p<2: return "≤1 peptide — PRM unreliable"
    if t1=="Singleton" and (pd.isna(row.get("Tier_IT_NT")) or t2=="Singleton"):
        return "Single-sample detection only"
    if row["Category"]=="Unclassified" and row["TotalScore"]<16:
        return "Unclassified, low score"
    return "Score below panel-size threshold"

candidates["TrimReason"] = candidates.apply(trim_reason, axis=1)

PANEL_COLS = [
    "Gene","Peptides","Category","CategoryPriority",
    "FC_IS_NS","BH_IS_NS","Tier_IS_NS","n_IS","n_NS",
    "FC_IT_NT","BH_IT_NT","Tier_IT_NT","n_IT","n_NT",
    "IPA_Pathways_ISNS","IPA_Pathways_ITNT",
    "TotalScore","PanelStatus","TrimReason"
]
selected_df = candidates[candidates["PanelStatus"]=="SELECTED"][PANEL_COLS].sort_values("TotalScore",ascending=False)
trimmed_df  = candidates[candidates["PanelStatus"]=="TRIMMED"][PANEL_COLS].sort_values("TotalScore",ascending=False)

# ── 13. IPA SUMMARY TABLES ────────────────────────────────────────────────────
sig_paths_out = all_paths[["label","Ingenuity Canonical Pathways","-log(p-value)","zScore","Ratio","Molecules","p_value"]].copy()
sig_paths_out.columns = ["Comparison","Pathway","neg_log10_p","zScore","Ratio","Molecules","p_value"]
sig_paths_out["Activated?"] = sig_paths_out["zScore"].apply(
    lambda z: "Activated" if (pd.notna(z) and z>=2) else
              "Inhibited" if (pd.notna(z) and z<=-2) else "n.d.")

ups_isns = isns_ups[["Upstream Regulator","Expr Log Ratio","Molecule Type",
                      "Predicted Activation State","Activation z-score",
                      "Bias-corrected z-score","p-value of overlap",
                      "Target molecules in dataset"]].copy()
ups_itnt = itnt_ups[["Upstream Regulator","Expr Log Ratio","Molecule Type",
                      "Predicted Activation State","Activation z-score",
                      "Bias-corrected z-score","p-value of overlap",
                      "Target molecules in dataset"]].copy()

# ── 14. EXCEL OUTPUT ──────────────────────────────────────────────────────────
wb = Workbook()

HDR  = PatternFill("solid", fgColor="1F3864")
SEL  = PatternFill("solid", fgColor="E2EFDA")
TRIM = PatternFill("solid", fgColor="FCE4D6")
TIER = {
    "BH<0.05":              (PatternFill("solid",fgColor="375623"), Font(color="FFFFFF",bold=True)),
    "BH<0.20":              (PatternFill("solid",fgColor="70AD47"), Font(color="FFFFFF")),
    "Present/Absent (n≥3)": (PatternFill("solid",fgColor="4472C4"), Font(color="FFFFFF",bold=True)),
    "Present/Absent (n=1-2)":(PatternFill("solid",fgColor="9DC3E6"),Font(color="000000")),
    "Trend":                (PatternFill("solid",fgColor="FFD966"), Font(color="000000")),
    "Singleton":            (PatternFill("solid",fgColor="FF0000"), Font(color="FFFFFF",bold=True)),
}
thin   = Side(style="thin",color="BFBFBF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)

COL_W = {
    "Gene":14,"Peptides":9,"Category":40,"CategoryPriority":10,
    "FC_IS_NS":11,"BH_IS_NS":11,"Tier_IS_NS":22,"n_IS":6,"n_NS":6,
    "FC_IT_NT":11,"BH_IT_NT":11,"Tier_IT_NT":22,"n_IT":6,"n_NT":6,
    "IPA_Pathways_ISNS":40,"IPA_Pathways_ITNT":40,
    "TotalScore":10,"PanelStatus":12,"TrimReason":40,
}

def write_gene_sheet(ws, df, title):
    ws.title = title
    hdrs = list(df.columns)
    tier_cols = {i+1 for i,h in enumerate(hdrs) if "Tier_" in h}
    status_col = next((i+1 for i,h in enumerate(hdrs) if h=="PanelStatus"),None)
    for ci,h in enumerate(hdrs,1):
        c = ws.cell(row=1,column=ci,value=h)
        c.font=Font(color="FFFFFF",bold=True,size=10); c.fill=HDR
        c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=border
    for ri,(_, row) in enumerate(df.iterrows(),2):
        st = row.get("PanelStatus","")
        rfill = SEL if st=="SELECTED" else TRIM
        for ci,h in enumerate(hdrs,1):
            val = row[h]
            if pd.isna(val): val=""
            if isinstance(val,float) and val==val: val=round(val,4)
            c=ws.cell(row=ri,column=ci,value=val)
            c.border=border; c.alignment=Alignment(wrap_text=False,vertical="center")
            if ci==1: c.font=Font(bold=True,size=10)
            else:     c.font=Font(size=9)
            if ci in tier_cols:
                tv=str(val)
                if tv in TIER: c.fill,c.font=TIER[tv]
                else:          c.fill=rfill
            else: c.fill=rfill
    for ci,h in enumerate(hdrs,1):
        ws.column_dimensions[get_column_letter(ci)].width=COL_W.get(h,12)
    ws.freeze_panes="B2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(hdrs))}{len(df)+1}"
    ws.row_dimensions[1].height=40
    ts_col=next((get_column_letter(i+1) for i,h in enumerate(hdrs) if h=="TotalScore"),None)
    if ts_col:
        ws.conditional_formatting.add(
            f"{ts_col}2:{ts_col}{len(df)+1}",
            ColorScaleRule(start_type="min",start_color="FFFFFF",
                           mid_type="percentile",mid_value=50,mid_color="FFEB84",
                           end_type="max",end_color="63BE7B"))

def write_table(ws, df, title, col_widths=None):
    ws.title=title
    hdrs=list(df.columns)
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=1,column=ci,value=h)
        c.font=Font(color="FFFFFF",bold=True); c.fill=HDR
        c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=border
    for ri,(_, row) in enumerate(df.iterrows(),2):
        for ci,h in enumerate(hdrs,1):
            val=row[h]
            if pd.isna(val): val=""
            if isinstance(val,float) and val==val: val=round(val,5)
            c=ws.cell(row=ri,column=ci,value=val)
            c.border=border; c.font=Font(size=9)
    ws.row_dimensions[1].height=35
    if col_widths:
        for ci,w in enumerate(col_widths,1):
            ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A2"

# Sheets
ws1=wb.active
write_gene_sheet(ws1,selected_df,"SELECTED Panel (≤80)")
ws2=wb.create_sheet("TRIMMED")
write_gene_sheet(ws2,trimmed_df,"TRIMMED")
ws3=wb.create_sheet("IPA Pathways (sig)")
write_table(ws3,sig_paths_out,"IPA Pathways (sig)",[12,55,14,12,10,60,12,15])
ws4=wb.create_sheet("Upstream Regs ISNS")
write_table(ws4,ups_isns.head(50),"Upstream Regs ISNS",[25,14,25,22,16,18,14,60])
ws5=wb.create_sheet("Upstream Regs ITNT")
write_table(ws5,ups_itnt.head(50),"Upstream Regs ITNT",[25,14,25,22,16,18,14,60])

# IPA Interpretation sheet
ws6=wb.create_sheet("IPA Interpretation")
rows=[
    ["BCC PRM Panel — IPA Integration Summary"],[""],
    ["ISNS STROMA COMPARISON (IS > NS) — KEY IPA FINDINGS"],
    ["Top pathway:","Collagen degradation (-log10p=6.68, z=+0.63)","Driven by MMP2, MMP12, COL4A1, COL6A3, CTSB, CTSK"],
    ["","Clathrin-mediated Endocytosis (7.28)","Highest pathway — integrin/vesicular trafficking enriched in IS stroma"],
    ["","RHO GTPase cycle (6.64, z=-0.6)","Cytoskeletal remodeling — inhibition direction suggests altered CAF contractility?"],
    ["","ECM Organization (5.42, z=-1.5)","INHIBITED — paradox: ECM proteins up but organising function down? Check z-score interpretation."],
    ["Top upstream:","TGFB1 — ACTIVATED (z=+2.6, p=2.6e-13)","★ STRONGEST FINDING: TGFβ1 is predicted activated in IS stroma vs NS. Directly supports hypothesis."],
    ["","TP63 — INHIBITED (z=-2.6, p=2.1e-12)","TP63 is a squamous/BCC identity TF. Its inhibition in IS stroma may reflect CAF-driven suppression of epithelial crosstalk."],
    ["","RYR1 (p=1.2e-16, no activation state)","Largest overlap but no z-score — may be a background hub gene in IPA. Do not over-interpret."],
    [""],
    ["ITNT TUMOR COMPARISON (IT > NT) — KEY IPA FINDINGS"],
    ["Top pathway:","Mitochondrial Translation (z=n.a., -log10p=16!)","16/97 mitochondrial ribosomal proteins. BIOLOGICAL ASSUMPTION FLAG: this likely reflects metabolic reprogramming in IT, not immune polarization directly."],
    ["","Processing of Capped Pre-mRNA (8.51, z=+4)","RNA splicing/processing activated in IT — possible link to transcriptional plasticity driving invasiveness?"],
    ["","TP53 Acetylation (6.80, z=+1.6)","EP300, CHD4 enriched — chromatin remodeling activated in IT."],
    ["","Keratinization (4.05, z=-2.12) — INHIBITED","KRT5, KRT14, KRT17, KRT20 DOWN in IT vs NT. Infiltrative BCC loses basal keratinocyte identity markers. These become NODULAR MARKERS for PRM."],
    ["","Elastic fibre formation (3.31, z=+1)","FBLN2, FBLN5, FN1 up in IT — confirms ECM remodeling in infiltrative compartment."],
    ["","Cytosolic DNA sensors (4.80, z=+1.6)","POLR2/3 subunits — innate immune sensing activated? Possible link to anti-tumor response or cGAS-STING pathway."],
    ["Top upstream:","EGFR (p=7.2e-8, no activation state)","Broad hub overlap in IT vs NT. Targets include KRT17, FN1, THBS1, LRRC15. Note: EGFR is a therapeutic target in resistant BCC."],
    ["","ARID1A — ACTIVATED (z=+2.5, p=1.1e-5)","Chromatin remodeling. Targets: FN1, KRT5, KRT14, THBS1. May reflect epigenetic reprogramming in infiltrative growth."],
    ["","ESR2 — ACTIVATED (z=+2.1, p=2.2e-5)","Estrogen receptor β in BCC is unexpected. FLAG: this may be a stromal cell (immune/endothelial) signal, not tumor-intrinsic. Confirm by IHC co-localization."],
    ["","TGFB1 (p=2.2e-4)","Also appears as upstream regulator in IT vs NT. No predicted activation state here (underpowered), but directionally consistent with ISNS finding."],
    ["","ARG1 (p=8.9e-5)","ARG1 appears as upstream regulator in IT vs NT with target genes GLS, IDH1, NNMT, FN1 — metabolic regulation."],
    [""],
    ["KEY ASSUMPTIONS TO FLAG IN MANUSCRIPT"],
    ["1.","Mitochondrial translation is the top IT vs NT pathway — this must be acknowledged as unexpected. Is it biology or a confounder (e.g., stromal cell contamination differences between IT and NT)?"],
    ["2.","TGFB1 activation (ISNS) is the strongest hypothesis-confirming finding but the actual TGFB1 protein shows BH=0.86 in ISNS. IPA infers activation from TARGET genes, not from TGFB1 protein level itself. Important distinction."],
    ["3.","ECM Organization is INHIBITED in ISNS (z=-1.5) despite individual collagen proteins being elevated. Possible explanation: the organizing enzymes/crosslinkers are down while raw collagen is up — dysfunctional matrix?"],
    ["4.","KRT5/14/17 being DOWN in IT vs NT is actually good news for the PRM panel: they become specific NODULAR markers, not infiltrative ones."],
    ["5.","ESR2 activation in IT vs NT is uninterpretable without knowing which cells express it. Before putting ESR2 or its targets on the panel for this reason, confirm by IHC."],
]
for ri,row in enumerate(rows,1):
    for ci,val in enumerate(row,1):
        c=ws6.cell(row=ri,column=ci,value=val)
        if ri==1: c.font=Font(bold=True,size=13,color="1F3864")
        elif ci==1 and val and val[0] in "ABCDEFGHIJKLMNOPQRSTUVWXYZTop": 
            c.font=Font(bold=True,size=10,color="1F3864")
            if len(row)>1: c.fill=PatternFill("solid",fgColor="D9E1F2")
        elif ci==2 and val and "★" in str(val):
            c.font=Font(bold=True,size=9,color="375623")
        elif ci==2 and "FLAG" in str(val).upper():
            c.font=Font(bold=True,size=9,color="FF0000")
        else: c.font=Font(size=9)
ws6.column_dimensions["A"].width=12
ws6.column_dimensions["B"].width=60
ws6.column_dimensions["C"].width=75

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Selected: {len(selected_df)} | Trimmed: {len(trimmed_df)}")
print("\nCategory coverage:")
for cat in CATEGORY_MIN:
    genes=[g for g in selected if cat in assign_categories(g)]
    print(f"  {cat}: {len(genes)}")
print("\nTop 15 in panel:")
print(selected_df[["Gene","Category","TotalScore","Tier_IS_NS","Tier_IT_NT",
                    "FC_IS_NS","FC_IT_NT","IPA_Pathways_ITNT"]].head(15).to_string(index=False))
